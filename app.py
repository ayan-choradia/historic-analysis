"""
US Fixed Income Analysis Dashboard
Analyzes meeting premiums (FED1-FED12), economic data, and Fed events from 2021-2025.
"""

import os
import json
import math
from datetime import datetime, timedelta
from flask import Flask, render_template, jsonify, request
import pandas as pd
import numpy as np
import openpyxl

app = Flask(__name__)

# ─── DATA PATHS ───
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ECON_CSV = os.path.join(BASE_DIR, 'data', 'US_Economic_Calendar_2021_2025_Clean.csv')
PREMIUMS_XLSX = os.path.join(BASE_DIR, 'data', 'US Meeting Premiums Generic Terms Since 2021.xlsx')

# ─── GLOBAL DATA STORE ───
DATA = {}


def safe_float(val):
    """Convert value to float, returning NaN for non-numeric."""
    if val is None or val == '' or val == 'NA' or val == 'na':
        return np.nan
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(',', '')
    # Handle percentages
    if s.endswith('%'):
        try:
            return float(s[:-1])
        except:
            return np.nan
    # Handle K/M/B suffixes
    multipliers = {'K': 1e3, 'M': 1e6, 'B': 1e9}
    for suffix, mult in multipliers.items():
        if s.endswith(suffix):
            try:
                return float(s[:-1]) * mult
            except:
                return np.nan
    try:
        return float(s)
    except:
        return np.nan


def classify_event(event_name):
    """Classify an economic event into a category."""
    e = event_name.lower()
    if 'fed interest rate' in e or 'fomc statement' in e:
        return 'FOMC Decision'
    if 'fomc economic projections' in e or 'fomc press conference' in e:
        return 'FOMC'
    if 'fomc meeting minutes' in e:
        return 'FOMC Minutes'
    if 'fed chair powell' in e or 'fomc member powell' in e:
        return 'Powell Speech'
    if 'fed chair' in e or 'fed monetary policy' in e:
        return 'Fed Speech'
    if 'kashkari' in e or 'waller' in e or 'bowman' in e or 'bullard' in e:
        return 'Fed Speech'
    if 'cpi (yoy)' in e or 'cpi (mom)' in e or 'core cpi' in e:
        return 'CPI'
    if 'pce price' in e or 'core pce' in e:
        return 'PCE'
    if 'nonfarm payrolls' in e:
        return 'NFP'
    if 'unemployment rate' in e:
        return 'Unemployment'
    if 'average hourly earnings' in e:
        return 'Wages'
    if 'initial jobless claims' in e:
        return 'Jobless Claims'
    if 'ism manufacturing pmi' in e:
        return 'ISM Mfg'
    if 'ism non-manufacturing pmi' in e or 'ism non-manufacturing' in e:
        return 'ISM Svc'
    if 'ism manufacturing prices' in e or 'ism non-manufacturing prices' in e:
        return 'ISM Prices'
    if 'retail sales' in e:
        return 'Retail Sales'
    if 'gdp' in e:
        return 'GDP'
    if 'ppi' in e:
        return 'PPI'
    if 'consumer confidence' in e:
        return 'Confidence'
    if 'durable goods' in e:
        return 'Durable Goods'
    if 'jolts' in e:
        return 'JOLTS'
    if 'adp nonfarm' in e:
        return 'ADP'
    if 'philadelphia fed' in e:
        return 'Philly Fed'
    if 'chicago pmi' in e:
        return 'Chicago PMI'
    if 'new home sales' in e or 'existing home sales' in e:
        return 'Housing'
    if '10-year note auction' in e or '30-year bond auction' in e:
        return 'Auction'
    if 'crude oil' in e:
        return 'Oil'
    if 'biden' in e or 'trump' in e or 'president' in e:
        return 'Political'
    if 'holiday' in e or 'day' in e.split(',')[-1] if ',' in e else False:
        return 'Holiday'
    return 'Other'


def get_surprise_direction(actual, consensus, event_name):
    """
    Determine if a data release was hawkish or dovish for rates.
    Returns: 1 = hawkish (higher rates), -1 = dovish (lower rates), 0 = inline
    """
    if pd.isna(actual) or pd.isna(consensus) or consensus == 0:
        return 0, 0.0

    surprise = actual - consensus

    if abs(surprise) < 1e-10:
        return 0, 0.0

    e = event_name.lower()

    # For inflation data: higher = hawkish
    if any(x in e for x in ['cpi', 'ppi', 'pce price', 'core pce', 'ism.*price',
                             'average hourly earnings']):
        return (1 if surprise > 0 else -1), surprise

    # For employment data: stronger = hawkish
    if any(x in e for x in ['nonfarm payrolls', 'adp nonfarm', 'jolts']):
        return (1 if surprise > 0 else -1), surprise

    # For unemployment/claims: higher = dovish (weaker economy)
    if any(x in e for x in ['unemployment rate', 'initial jobless claims']):
        return (-1 if surprise > 0 else 1), surprise

    # For activity/growth: stronger = hawkish
    if any(x in e for x in ['gdp', 'ism manufacturing pmi', 'ism non-manufacturing pmi',
                             'retail sales', 'consumer confidence', 'durable goods',
                             'philadelphia fed', 'chicago pmi', 'new home', 'existing home']):
        return (1 if surprise > 0 else -1), surprise

    return 0, surprise


def load_and_process_data():
    """Load all data sources and process them."""
    print("Loading meeting premiums...")
    wb = openpyxl.load_workbook(PREMIUMS_XLSX, data_only=True)
    ws = wb.active

    # Read all rows
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(row)

    # Build DataFrame
    df_prem = pd.DataFrame(rows, columns=headers)
    df_prem['date'] = pd.to_datetime(df_prem['Timestamp'])
    df_prem = df_prem.sort_values('date').reset_index(drop=True)

    # Convert FED columns to numeric
    fed_cols = [c for c in headers if c and c.startswith('FED')]
    for col in fed_cols:
        df_prem[col] = pd.to_numeric(df_prem[col], errors='coerce')

    # Calculate day-over-day changes
    for col in fed_cols:
        df_prem[f'{col}_chg'] = df_prem[col].diff()

    df_prem['date_str'] = df_prem['date'].dt.strftime('%Y-%m-%d')
    DATA['premiums'] = df_prem
    DATA['fed_cols'] = fed_cols
    DATA['fed_cols_12'] = [c for c in fed_cols if int(c.replace('FED', '')) <= 12]

    print(f"  Loaded {len(df_prem)} trading days of premiums, {len(fed_cols)} FED tenors")

    # ─── Economic Calendar ───
    # Builder df_econ and fill nan
    df_econ = pd.read_csv(ECON_CSV)
    df_econ = df_econ.replace({np.nan: None})
    df_econ['date'] = pd.to_datetime(df_econ['date'], format='mixed')
    df_econ['date_str'] = df_econ['date'].dt.strftime('%Y-%m-%d')

    # Parse numeric values
    df_econ['actual_num'] = df_econ['actual'].apply(safe_float)
    df_econ['consensus_num'] = df_econ['consensus'].apply(safe_float)
    df_econ['previous_num'] = df_econ['previous'].apply(safe_float)

    # Classify events
    df_econ['category'] = df_econ['event'].apply(classify_event)

    # Calculate surprise
    results = df_econ.apply(
        lambda r: get_surprise_direction(r['actual_num'], r['consensus_num'], r['event']),
        axis=1
    )
    df_econ['hawk_dove'] = [r[0] for r in results]  # 1=hawk, -1=dove, 0=inline
    df_econ['surprise_val'] = [r[1] for r in results]

    DATA['econ'] = df_econ
    print(f"  Loaded {len(df_econ)} economic events")

    # ─── Merge: daily premium changes with events ───
    print("Merging datasets...")
    chg_cols = [f'{c}_chg' for c in DATA['fed_cols_12']]

    # Daily summary of premium changes
    daily_prem = df_prem[['date_str'] + chg_cols + DATA['fed_cols_12']].copy()

    # Group events by date
    daily_events = df_econ.groupby('date_str').apply(
        lambda g: g[['event', 'actual', 'consensus', 'previous', 'category',
                      'hawk_dove', 'surprise_val', 'time']].to_dict('records')
    ).reset_index(name='events')

    merged = pd.merge(daily_prem, daily_events, on='date_str', how='left')
    DATA['merged'] = merged

    # ─── Precompute analytics ───
    print("Computing analytics...")

    # 1. Event-type impact analysis
    event_impacts = compute_event_impacts(df_econ, df_prem)
    DATA['event_impacts'] = event_impacts

    # 2. FOMC meeting dates and details
    fomc_dates = extract_fomc_dates(df_econ)
    DATA['fomc_dates'] = fomc_dates

    # 3. Regime classification
    DATA['regimes'] = classify_regimes(df_prem)

    # 4. Biggest moves
    DATA['biggest_moves'] = compute_biggest_moves(df_prem, df_econ)

    # 5. Terminal rate proxy evolution
    DATA['terminal_rate'] = compute_terminal_evolution(df_prem)

    print("Data processing complete!")


def compute_event_impacts(df_econ, df_prem):
    """Compute average premium moves by event category."""
    chg_cols_12 = [f'FED{i}_chg' for i in range(1, 13)]
    prem_daily = df_prem[['date_str'] + chg_cols_12].copy()

    # Merge events with premium changes
    events_with_prem = pd.merge(
        df_econ[['date_str', 'event', 'category', 'hawk_dove', 'surprise_val']],
        prem_daily,
        on='date_str',
        how='inner'
    )

    results = {}
    important_cats = ['CPI', 'NFP', 'PCE', 'FOMC Decision', 'FOMC Minutes',
                      'Powell Speech', 'Fed Speech', 'ISM Mfg', 'ISM Svc',
                      'GDP', 'PPI', 'Retail Sales', 'Jobless Claims', 'JOLTS',
                      'ADP', 'Wages', 'Unemployment', 'Confidence']

    for cat in important_cats:
        cat_data = events_with_prem[events_with_prem['category'] == cat]
        if len(cat_data) == 0:
            continue

        # Overall average absolute move
        avg_moves = {}
        avg_moves_signed = {}
        hawk_moves = {}
        dove_moves = {}

        for col in chg_cols_12:
            fed_name = col.replace('_chg', '')
            valid = cat_data[col].dropna()
            avg_moves[fed_name] = float(valid.abs().mean()) if len(valid) > 0 else 0
            avg_moves_signed[fed_name] = float(valid.mean()) if len(valid) > 0 else 0

            # Hawkish surprises
            hawk_data = cat_data[cat_data['hawk_dove'] == 1][col].dropna()
            hawk_moves[fed_name] = float(hawk_data.mean()) if len(hawk_data) > 0 else 0

            # Dovish surprises
            dove_data = cat_data[cat_data['hawk_dove'] == -1][col].dropna()
            dove_moves[fed_name] = float(dove_data.mean()) if len(dove_data) > 0 else 0

        results[cat] = {
            'count': int(len(cat_data)),
            'avg_abs_move': avg_moves,
            'avg_signed_move': avg_moves_signed,
            'hawk_avg_move': hawk_moves,
            'dove_avg_move': dove_moves,
            'hawk_count': int((cat_data['hawk_dove'] == 1).sum()),
            'dove_count': int((cat_data['hawk_dove'] == -1).sum()),
            'inline_count': int((cat_data['hawk_dove'] == 0).sum()),
        }

    return results


def extract_fomc_dates(df_econ):
    """Extract FOMC meeting dates and related info."""
    fomc_events = df_econ[
        df_econ['event'].str.contains('Fed Interest Rate Decision', case=False, na=False)
    ].copy()

    meetings = []
    for _, row in fomc_events.iterrows():
        date_str = row['date_str']
        actual = row['actual']
        consensus = row['consensus']
        previous = row['previous']

        # Check if there was a SEP
        same_day = df_econ[df_econ['date_str'] == date_str]
        has_sep = same_day['event'].str.contains('FOMC Economic Projections', case=False, na=False).any()

        meetings.append({
            'date': date_str,
            'actual': actual,
            'consensus': consensus,
            'previous': previous,
            'has_sep': bool(has_sep),
            'surprise': bool(actual != consensus if pd.notna(actual) and pd.notna(consensus) else False),
        })

    return meetings


def classify_regimes(df_prem):
    """Classify time periods into policy regimes."""
    regimes = [
        {'name': 'Zero Rates & Transitory', 'start': '2021-01-01', 'end': '2021-11-30',
         'color': '#4CAF50', 'description': 'Fed at zero, inflation called transitory, QE ongoing'},
        {'name': 'Taper & Hawkish Pivot', 'start': '2021-12-01', 'end': '2022-03-15',
         'color': '#FF9800', 'description': 'Transitory retired, taper accelerated, liftoff March 2022'},
        {'name': 'Aggressive Hiking', 'start': '2022-03-16', 'end': '2022-12-31',
         'color': '#F44336', 'description': '425bps in 9 months, 75bp mega-hikes, Jackson Hole pain speech'},
        {'name': 'Slower Hikes + Banking Crisis', 'start': '2023-01-01', 'end': '2023-05-31',
         'color': '#E91E63', 'description': '25bp hikes, SVB collapse, banking turmoil'},
        {'name': 'Higher For Longer', 'start': '2023-06-01', 'end': '2023-11-27',
         'color': '#9C27B0', 'description': 'Skip-hike-hold pattern, rates at 5.25-5.50%, 10Y hit 5%'},
        {'name': 'Dovish Pivot & Cuts Anticipation', 'start': '2023-11-28', 'end': '2024-09-17',
         'color': '#2196F3', 'description': 'Waller pivot speech, Dec 2023 dovish dots, market pricing 6-7 cuts'},
        {'name': 'Rate Cutting Cycle', 'start': '2024-09-18', 'end': '2024-12-31',
         'color': '#00BCD4', 'description': '50bp cut Sep, 25bp Nov & Dec, 100bp total. Dec hawkish dot surprise'},
        {'name': 'Tariff Uncertainty & Pause', 'start': '2025-01-01', 'end': '2025-12-31',
         'color': '#607D8B', 'description': 'Trump tariffs, extended pause at 4.25-4.50%, Liberation Day'},
    ]

    # Calculate stats for each regime
    for regime in regimes:
        mask = (df_prem['date_str'] >= regime['start']) & (df_prem['date_str'] <= regime['end'])
        regime_data = df_prem[mask]

        if len(regime_data) > 0:
            regime['trading_days'] = int(len(regime_data))
            # Average premium levels
            regime['avg_premiums'] = {}
            for i in range(1, 13):
                col = f'FED{i}'
                if col in regime_data.columns:
                    vals = regime_data[col].dropna()
                    regime['avg_premiums'][col] = round(float(vals.mean()), 2) if len(vals) > 0 else None

            # Average daily volatility (abs change)
            regime['avg_volatility'] = {}
            for i in range(1, 13):
                col = f'FED{i}_chg'
                if col in regime_data.columns:
                    vals = regime_data[col].dropna()
                    regime['avg_volatility'][f'FED{i}'] = round(float(vals.abs().mean()), 2) if len(vals) > 0 else None

            # Start and end premium levels
            first_valid = regime_data.iloc[0] if len(regime_data) > 0 else None
            last_valid = regime_data.iloc[-1] if len(regime_data) > 0 else None
            regime['start_premiums'] = {}
            regime['end_premiums'] = {}
            for i in range(1, 13):
                col = f'FED{i}'
                if col in regime_data.columns:
                    regime['start_premiums'][col] = round(float(first_valid[col]), 2) if pd.notna(first_valid[col]) else None
                    regime['end_premiums'][col] = round(float(last_valid[col]), 2) if pd.notna(last_valid[col]) else None
        else:
            regime['trading_days'] = 0
            regime['avg_premiums'] = {}
            regime['avg_volatility'] = {}

    return regimes


def compute_biggest_moves(df_prem, df_econ):
    """Find the days with the biggest meeting premium moves."""
    # Sum of absolute changes across FED1-FED8 as a measure of total curve move
    chg_cols = [f'FED{i}_chg' for i in range(1, 9)]
    valid_cols = [c for c in chg_cols if c in df_prem.columns]

    df = df_prem.copy()
    df['total_abs_move'] = df[valid_cols].abs().sum(axis=1)
    df['total_signed_move'] = df[valid_cols].sum(axis=1)
    df['front_move'] = df[[f'FED{i}_chg' for i in range(1, 4) if f'FED{i}_chg' in df.columns]].sum(axis=1)
    df['back_move'] = df[[f'FED{i}_chg' for i in range(5, 9) if f'FED{i}_chg' in df.columns]].sum(axis=1)

    # Top 50 biggest absolute move days
    top_moves = df.nlargest(50, 'total_abs_move')

    results = []
    for _, row in top_moves.iterrows():
        date_str = row['date_str']
        events = df_econ[df_econ['date_str'] == date_str][['event', 'actual', 'consensus', 'category']].to_dict('records')

        results.append({
            'date': date_str,
            'total_abs_move': round(float(row['total_abs_move']), 2),
            'total_signed_move': round(float(row['total_signed_move']), 2),
            'front_move': round(float(row['front_move']), 2),
            'back_move': round(float(row['back_move']), 2),
            'fed1_chg': round(float(row.get('FED1_chg', 0)), 2) if pd.notna(row.get('FED1_chg')) else 0,
            'fed2_chg': round(float(row.get('FED2_chg', 0)), 2) if pd.notna(row.get('FED2_chg')) else 0,
            'fed3_chg': round(float(row.get('FED3_chg', 0)), 2) if pd.notna(row.get('FED3_chg')) else 0,
            'fed4_chg': round(float(row.get('FED4_chg', 0)), 2) if pd.notna(row.get('FED4_chg')) else 0,
            'events': events[:5],  # Top 5 events
        })

    return results


def compute_terminal_evolution(df_prem):
    """Track implied terminal rate proxy over time using the furthest valid FED meeting premium."""
    results = []
    for _, row in df_prem.iterrows():
        date_str = row['date_str']
        # Use FED8 as a proxy for terminal rate implied (8 meetings out ≈ 2 years)
        # The cumulative premium tells us the total expected rate change
        cumulative = 0
        for i in range(1, 9):
            col = f'FED{i}'
            if col in row and pd.notna(row[col]):
                cumulative += row[col]

        results.append({
            'date': date_str,
            'cumulative_8': round(cumulative, 2),
            'fed1': round(float(row.get('FED1', 0)), 2) if pd.notna(row.get('FED1')) else None,
            'fed4': round(float(row.get('FED4', 0)), 2) if pd.notna(row.get('FED4')) else None,
            'fed8': round(float(row.get('FED8', 0)), 2) if pd.notna(row.get('FED8')) else None,
        })

    return results


def json_safe(obj):
    """Make object JSON serializable."""
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating, float)):
        if pd.isna(obj) or (isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj))):
            return None
        return float(obj)
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, pd.Timestamp):
        return obj.strftime('%Y-%m-%d')
    if isinstance(obj, datetime):
        return obj.strftime('%Y-%m-%d')
    return obj


class SafeEncoder(json.JSONEncoder):
    def default(self, obj):
        result = json_safe(obj)
        if result is not obj:
            return result
        return super().default(obj)


# ─── ROUTES ───

@app.route('/')
def dashboard():
    return render_template('dashboard.html')


@app.route('/day-explorer')
def day_explorer():
    return render_template('day_explorer.html')


@app.route('/curve-viewer')
def curve_viewer():
    return render_template('curve_viewer.html')


@app.route('/event-analysis')
def event_analysis():
    return render_template('event_analysis.html')


@app.route('/fomc-tracker')
def fomc_tracker():
    return render_template('fomc_tracker.html')


@app.route('/regime-analysis')
def regime_analysis():
    return render_template('regime_analysis.html')


# ─── API ROUTES ───

@app.route('/api/dashboard-stats')
def api_dashboard_stats():
    df = DATA['premiums']
    chg_cols = [f'FED{i}_chg' for i in range(1, 9)]

    stats = {
        'total_trading_days': int(len(df)),
        'date_range': f"{df['date_str'].iloc[0]} to {df['date_str'].iloc[-1]}",
        'total_events': int(len(DATA['econ'])),
        'fomc_meetings': len(DATA['fomc_dates']),
        'biggest_single_day_move': DATA['biggest_moves'][0] if DATA['biggest_moves'] else None,
        'avg_daily_vol': {
            f'FED{i}': round(float(df[f'FED{i}_chg'].dropna().abs().mean()), 2)
            for i in range(1, 13)
            if f'FED{i}_chg' in df.columns
        },
        'current_premiums': {
            f'FED{i}': round(float(df[f'FED{i}'].iloc[-1]), 2)
            for i in range(1, 13)
            if f'FED{i}' in df.columns and pd.notna(df[f'FED{i}'].iloc[-1])
        },
    }
    return json.dumps(stats, cls=SafeEncoder)


@app.route('/api/premium-timeseries')
def api_premium_timeseries():
    df = DATA['premiums']
    # Downsample for performance: every nth day
    n = max(1, len(df) // 500)
    sampled = df.iloc[::n]

    data = {
        'dates': sampled['date_str'].tolist(),
    }
    for i in range(1, 13):
        col = f'FED{i}'
        if col in sampled.columns:
            data[col] = [round(float(v), 2) if pd.notna(v) else None for v in sampled[col]]

    return json.dumps(data, cls=SafeEncoder)


@app.route('/api/premium-changes-timeseries')
def api_premium_changes_ts():
    df = DATA['premiums']
    n = max(1, len(df) // 500)
    sampled = df.iloc[::n]

    data = {
        'dates': sampled['date_str'].tolist(),
    }
    for i in range(1, 13):
        col = f'FED{i}_chg'
        if col in sampled.columns:
            data[f'FED{i}'] = [round(float(v), 2) if pd.notna(v) else None for v in sampled[col]]

    return json.dumps(data, cls=SafeEncoder)


@app.route('/api/day-data/<date_str>')
def api_day_data(date_str):
    df_prem = DATA['premiums']
    df_econ = DATA['econ']

    # Get premium data for this date
    prem_row = df_prem[df_prem['date_str'] == date_str]
    if len(prem_row) == 0:
        return jsonify({'error': 'No premium data for this date'})

    prem_row = prem_row.iloc[0]

    # Get events
    events = df_econ[df_econ['date_str'] == date_str].to_dict('records')
    for ev in events:
        for k, v in ev.items():
            ev[k] = json_safe(v)
            if isinstance(ev[k], float) and (math.isnan(ev[k]) or math.isinf(ev[k])):
                ev[k] = None

    # Premium levels and changes
    premiums = {}
    changes = {}
    for i in range(1, 13):
        col = f'FED{i}'
        chg_col = f'{col}_chg'
        premiums[col] = round(float(prem_row[col]), 2) if pd.notna(prem_row.get(col)) else None
        changes[col] = round(float(prem_row[chg_col]), 2) if pd.notna(prem_row.get(chg_col)) else None

    # Previous day data for curve comparison
    idx = prem_row.name
    prev_premiums = {}
    if idx > 0:
        prev_row = df_prem.iloc[idx - 1]
        for i in range(1, 13):
            col = f'FED{i}'
            prev_premiums[col] = round(float(prev_row[col]), 2) if pd.notna(prev_row.get(col)) else None

    result = {
        'date': date_str,
        'premiums': premiums,
        'changes': changes,
        'prev_premiums': prev_premiums,
        'events': events,
        'total_front_change': sum(v for v in [changes.get(f'FED{i}') for i in range(1, 4)] if v is not None),
        'total_back_change': sum(v for v in [changes.get(f'FED{i}') for i in range(5, 9)] if v is not None),
    }

    return json.dumps(result, cls=SafeEncoder)


@app.route('/api/available-dates')
def api_available_dates():
    dates = DATA['premiums']['date_str'].tolist()
    return jsonify(dates)


@app.route('/api/event-impacts')
def api_event_impacts():
    return json.dumps(DATA['event_impacts'], cls=SafeEncoder)


@app.route('/api/fomc-meetings')
def api_fomc_meetings():
    meetings = DATA['fomc_dates']
    df_prem = DATA['premiums']

    enriched = []
    for m in meetings:
        date_str = m['date']
        prem_row = df_prem[df_prem['date_str'] == date_str]

        changes = {}
        premiums = {}
        if len(prem_row) > 0:
            row = prem_row.iloc[0]
            for i in range(1, 13):
                col = f'FED{i}'
                chg = f'{col}_chg'
                changes[col] = round(float(row[chg]), 2) if pd.notna(row.get(chg)) else None
                premiums[col] = round(float(row[col]), 2) if pd.notna(row.get(col)) else None

        # Get 5-day window of changes
        window = []
        for offset in range(-2, 3):
            try:
                idx = df_prem[df_prem['date_str'] == date_str].index[0] - offset
                if 0 <= idx < len(df_prem):
                    w_row = df_prem.iloc[idx]
                    w_changes = {}
                    for i in range(1, 9):
                        chg = f'FED{i}_chg'
                        w_changes[f'FED{i}'] = round(float(w_row[chg]), 2) if pd.notna(w_row.get(chg)) else None
                    window.append({
                        'date': w_row['date_str'],
                        'offset': offset,
                        'changes': w_changes,
                    })
            except:
                pass

        enriched.append({
            **m,
            'changes': changes,
            'premiums': premiums,
            'window': window,
        })

    return json.dumps(enriched, cls=SafeEncoder)


@app.route('/api/regimes')
def api_regimes():
    return json.dumps(DATA['regimes'], cls=SafeEncoder)


@app.route('/api/biggest-moves')
def api_biggest_moves():
    return json.dumps(DATA['biggest_moves'], cls=SafeEncoder)


@app.route('/api/terminal-rate')
def api_terminal_rate():
    data = DATA['terminal_rate']
    # Downsample
    n = max(1, len(data) // 500)
    sampled = data[::n]
    return json.dumps(sampled, cls=SafeEncoder)


@app.route('/api/curve-snapshot/<date_str>')
def api_curve_snapshot(date_str):
    df = DATA['premiums']
    row = df[df['date_str'] == date_str]
    if len(row) == 0:
        return jsonify({'error': 'Date not found'})

    row = row.iloc[0]
    curve = {}
    for i in range(1, 21):
        col = f'FED{i}'
        if col in row.index and pd.notna(row[col]):
            curve[col] = round(float(row[col]), 2)

    return json.dumps({'date': date_str, 'curve': curve}, cls=SafeEncoder)


@app.route('/api/search-events')
def api_search_events():
    query = request.args.get('q', '').lower()
    cat = request.args.get('category', '')

    df = DATA['econ']
    mask = pd.Series([True] * len(df))

    if query:
        mask = mask & df['event'].str.lower().str.contains(query, na=False)
    if cat:
        mask = mask & (df['category'] == cat)

    results = df[mask].head(200)
    records = []
    for _, row in results.iterrows():
        r = {
            'date': row['date_str'],
            'event': row['event'],
            'actual': row['actual'],
            'consensus': row['consensus'],
            'previous': row['previous'],
            'category': row['category'],
            'hawk_dove': int(row['hawk_dove']),
        }
        records.append(r)

    return json.dumps(records, cls=SafeEncoder)


# Initialize data for Vercel Serverless
try:
    load_and_process_data()
except Exception as e:
    print("Initialization error:", e)

if __name__ == '__main__':
    app.run(debug=True, port=5050)
